import os
import os.path
import pandas as pd
import parser
import cleanup
from openpyxl import load_workbook


def createMembersSheet(arr, fileHostname):
    
    membersData = arr[0]

    writer = pd.ExcelWriter(f'ExcelSheetOutputs/{fileHostname}.xlsx', engine='xlsxwriter')

    df = pd.DataFrame(membersData)
    df.to_excel(writer, sheet_name='Stack', index=False, header=False)
    writer.save()




def createPowerSupplySheet(arr, fileSavePath):

    powerSupplyData = arr[1]

    book = load_workbook(fileSavePath)        
    writer = pd.ExcelWriter(fileSavePath, engine='openpyxl')
    writer.book = book

    if 'Power-Supply' in book.sheetnames:
        book.remove(book['Power-Supply'])


    df = pd.DataFrame(powerSupplyData)
    df.to_excel(writer, sheet_name='Power-Supply', index=False, header=False)
    writer.save()



def createDacSheet(arr, fileSavePath):

    dacData = arr[2]

    book = load_workbook(fileSavePath)        
    writer = pd.ExcelWriter(fileSavePath, engine='openpyxl')
    writer.book = book

    if 'DAC' in book.sheetnames:
        book.remove(book['DAC'])


    df = pd.DataFrame(dacData)
    df.to_excel(writer, sheet_name='DAC', index=False, header=False)
    writer.save()


def createUplinkModuleSheet(arr, fileSavePath):
    uplinkModuleData = arr[3]

    book = load_workbook(fileSavePath)        
    writer = pd.ExcelWriter(fileSavePath, engine='openpyxl')
    writer.book = book

    if 'Uplink Module' in book.sheetnames:
        book.remove(book['Uplink Module'])


    df = pd.DataFrame(uplinkModuleData)
    df.to_excel(writer, sheet_name='Uplink Module', index=False, header=False)
    writer.save()


# def createSfpSheet(rawDataArray, startPoint, endPoint, linesUntilDataStart, fileSavePath):

#     desiredArray = parser.createDesiredArray(rawDataArray, startPoint, endPoint)

#     if desiredArray:
#         dataOnlyArray = parser.createDataOnlyArray(desiredArray, linesUntilDataStart)
#     else:
#         dataOnlyArray = ['No SFPs in Stack']
    
#     writer = pd.ExcelWriter(fileSavePath, engine='xlsxwriter')

#     df = pd.DataFrame(dataOnlyArray)
#     df.to_excel(writer, sheet_name='SFPs', index=False, header=False)
#     writer.save()

#     return dataOnlyArray



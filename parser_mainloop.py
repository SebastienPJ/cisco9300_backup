import os
import parser
import create_sheets
# import object
import pandas as pd
from openpyxl import load_workbook


def parseConfigFile(filepath):

    unfilteredArray = parser.createUnfilteredArray(filepath)
    hostname = parser.getHostname(unfilteredArray)
    sheetPath = f'ExcelSheetOutputs/{hostname}.xlsx'


    """INVENTORY DATA"""
    dataBeginPoint = [f'{hostname}#', 'show', 'inventory']
    dataEndPoint = [f'{hostname}#', '!!!!!!!!!!']
    dataLinesUntilActualDataStart = 1

    stackArray = parser.createStackArray(unfilteredArray, dataBeginPoint, dataEndPoint, dataLinesUntilActualDataStart)

    create_sheets.createMembersSheet(stackArray, hostname)

    create_sheets.createPowerSupplySheet(stackArray, sheetPath)

    create_sheets.createDacSheet(stackArray, sheetPath)

    if stackArray[3]:  # Uplink Module info is on stackArray[3], checks if there is any info there
        create_sheets.createUplinkModuleSheet(stackArray, sheetPath)


    # """""SFP"""
    # SfpBeginPoint = ['Port', 'Type', 'Product', 'Serial', 'Part']
    # SfpEndPoint =  f'{hostname}#'
    # SfpLinesUntilActualDataStart = 3 

  

    # sfpData = create_sheets.createSfpSheet(unfilteredArray, SfpBeginPoint, SfpEndPoint, SfpLinesUntilActualDataStart, sheetPath)  



    # # os.system(f'start EXCEL.EXE {sheetName}')
    # # switchStackObj = object.generateStackObject(sfpData, membersData, powerSupplyData)
    # # print(switchStackObj['member1']['device'])
    # # print(switchStackObj)



    # # if 'StackObj' in book.sheetnames:
    # #     book.remove(book['StackObj'])


    # # df = pd.DataFrame(switchStackObj)
    # # df.to_excel(writer, sheet_name='StackObj')

    # # print(sfpData)


